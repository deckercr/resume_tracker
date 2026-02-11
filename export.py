import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def export_to_excel(applications, job_descriptions):
    wb = Workbook()

    # --- Sheet 1: Applications ---
    ws1 = wb.active
    ws1.title = 'Applications'
    headers = [
        'Date', 'Status', 'Company', 'Position', 'Location',
        'URL', 'Job ID', 'Salary', 'Contact', 'Notes',
    ]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = Font(bold=True)

    for app in applications:
        ws1.append([
            app.date,
            app.status,
            app.company,
            app.position,
            app.location,
            app.url,
            app.job_id,
            app.salary,
            app.contact,
            app.notes,
        ])

    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = 'YYYY-MM-DD'

    for col in ws1.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # --- Sheet 2: Job Description ---
    ws2 = wb.create_sheet('Job Description')
    ws2.append(['Company', 'Position', 'Job Description'])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for jd in job_descriptions:
        ws2.append([jd.company, jd.position, jd.description])

    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 80

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    tmp.close()
    return tmp.name
